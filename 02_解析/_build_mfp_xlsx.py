#!/usr/bin/env python3
"""昨年度9班データから空気中の平均自由行程を計算する Excel を生成する。

昨年と同じ定義（熱中性子・管理棟1階→白根山）:
  空気を一様密度 ρ = 1.00×10⁻³ g/cm³ とする
  Δh ≈ 2000 m（白根山を約2000 m、KEKを0 m と近似）
  λ = Δh / ln(φ₂/φ₁)  =  Λ / ρ
  Λ = ρ Δh / ln(φ₂/φ₁)

  φ₁ = 1.83×10⁻³, φ₂ = 7.10×10⁻³
  → ln(φ₂/φ₁) ≈ 1.356,  λ ≈ 1470 m,  Λ ≈ 147 g/cm²
"""

from __future__ import annotations

import math

from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

BLACK, GRAY, BLUE, RED = "000000", "666666", "1F77B4", "D62728"
PALE, WHITE, HEADER, YELLOW = "F5F5F5", "FFFFFF", "1F4E79", "FFF2CC"

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

RHO_AIR = 1.00e-3  # g/cm³
DH_SHIRANE = 2000.0  # 昨年が使った標高差 [m]
H_TSUKUBA = 877.0
PHI_TH_1F, PHI_TH_TSU, PHI_TH_SHI = 1.83e-3, 3.12e-3, 7.10e-3
PHI_MEV_1F, PHI_MEV_TSU, PHI_MEV_SHI = 7.25e-4, 1.89e-3, 5.37e-3


def lam_uniform(dh, p1, p2):
    ratio = p2 / p1
    lam_m = dh / math.log(ratio)
    Lam = RHO_AIR * dh * 100.0 / math.log(ratio)  # g/cm²
    return ratio, Lam, lam_m


def setup_axis(axis, title, major, minor=None, nfmt="General", amin=None, amax=None):
    axis.title = title
    axis.delete = False
    axis.majorTickMark = "out"
    axis.minorTickMark = "out" if minor else "none"
    axis.tickLblPos = "nextTo"
    axis.majorUnit = major
    if minor:
        axis.minorUnit = minor
    axis.numFmt = nfmt
    if amin is not None:
        axis.scaling.min = amin
    if amax is not None:
        axis.scaling.max = amax
    axis.majorGridlines = ChartLines(
        spPr=GraphicalProperties(ln=LineProperties(solidFill="D9D9D9", w=9525))
    )
    if minor:
        axis.minorGridlines = ChartLines(
            spPr=GraphicalProperties(ln=LineProperties(solidFill="EFEFEF", w=6350))
        )
    axis.spPr = GraphicalProperties(ln=LineProperties(solidFill=BLACK, w=15875))


def header(ws, row, n, start=1):
    for c in range(start, start + n):
        cell = ws.cell(row, c)
        cell.font = Font(bold=True, color=WHITE, size=11)
        cell.fill = PatternFill("solid", fgColor=HEADER)
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[row].height = 28


def box(ws, r, c, val=None, fmt=None, fill=None, bold=False, align=center):
    cell = ws.cell(r, c, val)
    cell.border = border
    cell.alignment = align
    cell.font = Font(size=11, bold=bold)
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    return cell


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "平均自由行程"

    th_ratio, th_Lam, th_lam = lam_uniform(DH_SHIRANE, PHI_TH_1F, PHI_TH_SHI)
    mev_ratio, mev_Lam, mev_lam = lam_uniform(DH_SHIRANE, PHI_MEV_1F, PHI_MEV_SHI)

    ws["B2"] = "宇宙線中性子の平均自由行程"
    ws["B2"].font = Font(size=20, bold=True, color=HEADER)
    ws["B3"] = "昨年度 KEKサマーチャレンジ9班（熱中性子・管理棟1階基準）"
    ws["B3"].font = Font(size=11, color=GRAY)

    ws["B5"] = "計算式（昨年と同じ）"
    ws["B5"].font = Font(size=13, bold=True)
    ws["B6"] = "λ = Δh / ln(φ₂/φ₁)　　［m］"
    ws["B7"] = "Λ = ρ_air Δh / ln(φ₂/φ₁)　　［g/cm²］　　（ρ_air = 1.00×10⁻³ g/cm³ → λ [m] = 10 × Λ）"
    ws["B8"] = "空気は一様密度。標高差 Δh = 2000 m（白根山≈2000 m、KEK≈0 m）。標準大気の X(h) は使わない。"
    for r in (6, 7):
        ws.cell(r, 2).font = Font(size=13, bold=True)
    ws["B8"].font = Font(size=10, color=GRAY)

    ws["B10"] = "使用データ"
    ws["B10"].font = Font(size=13, bold=True)
    for c, h in enumerate(
        ["地点", "標高差 Δh (m)", "熱中性子 φ (×10⁻³)", "MeV φ (×10⁻³)"], 2
    ):
        ws.cell(11, c, h)
    header(ws, 11, 4, start=2)

    data = [
        ("管理棟1階（基準）", 0.0, PHI_TH_1F, PHI_MEV_1F),
        ("筑波山", H_TSUKUBA, PHI_TH_TSU, PHI_MEV_TSU),
        ("白根山", DH_SHIRANE, PHI_TH_SHI, PHI_MEV_SHI),
    ]
    for i, (name, h, th, mev) in enumerate(data, 12):
        box(ws, i, 2, name, align=left)
        box(ws, i, 3, h, "#,##0")
        box(ws, i, 4, th * 1e3, "0.00")
        box(ws, i, 5, mev * 1e3, "0.00")
    ws["B15"] = (
        "フラックスは公開表の値。標高差は昨年どおり白根山を 2000 m と近似。"
        "地理的な標高は KEK≈30 m、草津白根≈2160 m。"
    )
    ws["B15"].font = Font(size=10, color=GRAY)

    ws["B17"] = "計算結果（管理棟1階 → 白根山、Δh = 2000 m）"
    ws["B17"].font = Font(size=13, bold=True)
    for c, h in enumerate(["領域", "φ₂/φ₁", "Λ (g/cm²)", "λ (m)", "備考"], 2):
        ws.cell(18, c, h)
    header(ws, 18, 5, start=2)

    box(ws, 19, 2, "熱中性子", align=left, fill=YELLOW, bold=True)
    box(ws, 19, 3, th_ratio, "0.00", fill=YELLOW, bold=True)
    box(ws, 19, 4, th_Lam, "0.0", fill=YELLOW, bold=True)
    box(ws, 19, 5, round(th_lam), "#,##0", fill=YELLOW, bold=True)
    box(ws, 19, 6, "昨年の主結果", align=left, fill=YELLOW, bold=True)

    box(ws, 20, 2, "MeV領域", align=left)
    box(ws, 20, 3, mev_ratio, "0.00")
    box(ws, 20, 4, mev_Lam, "0.0")
    box(ws, 20, 5, round(mev_lam), "#,##0")
    box(ws, 20, 6, "同じ定義での参考", align=left)

    ws["B22"] = "主結果"
    ws["B22"].font = Font(size=12, bold=True)
    ws["C22"] = f"熱中性子の平均自由行程  λ = {th_lam:.0f} m"
    ws["C22"].font = Font(size=16, bold=True, color=BLUE)
    ws.merge_cells("C22:F22")
    ws["C23"] = (
        f"管理棟1階→白根山、Δh = 2000 m、Λ = {th_Lam:.0f} g/cm²、"
        "ρ_air = 1.00×10⁻³ g/cm³。昨年と同じ定義。"
    )
    ws["C23"].font = Font(size=10, color=GRAY)
    ws.merge_cells("C23:F23")
    ws["C24"] = (
        "1730 m になる計算法は、標準大気の大気深度 X(h) を使った場合。"
        "空気が高度とともに薄くなるため ΔX が大きくなり、λ も長めに出る。"
    )
    ws["C24"].font = Font(size=10, color=GRAY)
    ws.merge_cells("C24:F24")

    ws["B26"] = "グラフ用"
    ws["B26"].font = Font(size=13, bold=True)
    for c, h in enumerate(["標高差 (m)", "熱 φ (×10⁻³)", "MeV φ (×10⁻³)"], 2):
        ws.cell(27, c, h)
    header(ws, 27, 3, start=2)
    for i, (name, h, th, mev) in enumerate(data, 28):
        box(ws, i, 2, h, "#,##0")
        box(ws, i, 3, th * 1e3, "0.00")
        box(ws, i, 4, mev * 1e3, "0.00")

    ws["F27"] = "標高差 (m)"
    ws["G27"] = f"熱 予測 (λ={th_lam:.0f} m)"
    ws["H27"] = f"MeV 予測 (λ={mev_lam:.0f} m)"
    header(ws, 27, 3, start=6)
    heights = list(range(0, 2001, 200))
    for i, h in enumerate(heights, 28):
        th_fit = PHI_TH_1F * math.exp(h / th_lam) * 1e3
        mev_fit = PHI_MEV_1F * math.exp(h / mev_lam) * 1e3
        box(ws, i, 6, h, "#,##0")
        box(ws, i, 7, th_fit, "0.00")
        box(ws, i, 8, mev_fit, "0.00")
    fit_last = 27 + len(heights)

    chart = ScatterChart()
    chart.title = f"フラックスの高度依存（管理棟1階基準、熱中性子 λ = {th_lam:.0f} m）"
    chart.height = 12
    chart.width = 18
    chart.legend.position = "b"
    setup_axis(chart.x_axis, "標高差 Δh (m)", major=400, minor=200, nfmt="#,##0", amin=0, amax=2200)
    setup_axis(chart.y_axis, "フラックス (×10⁻³ s⁻¹ cm⁻²)", major=1, minor=0.5,
               nfmt="0.0", amin=0, amax=8)

    x_meas = Reference(ws, min_col=2, min_row=28, max_row=30)
    s_th = Series(Reference(ws, min_col=3, min_row=28, max_row=30), x_meas, title="熱中性子（測定）")
    s_th.graphicalProperties.line.noFill = True
    s_th.marker = Marker(symbol="circle", size=11)
    s_th.marker.graphicalProperties.solidFill = BLUE
    s_th.marker.graphicalProperties.line.solidFill = BLUE

    s_mev = Series(Reference(ws, min_col=4, min_row=28, max_row=30), x_meas, title="MeV（測定）")
    s_mev.graphicalProperties.line.noFill = True
    s_mev.marker = Marker(symbol="triangle", size=12)
    s_mev.marker.graphicalProperties.solidFill = RED
    s_mev.marker.graphicalProperties.line.solidFill = RED

    x_fit = Reference(ws, min_col=6, min_row=28, max_row=fit_last)
    s_th_fit = Series(Reference(ws, min_col=7, min_row=28, max_row=fit_last), x_fit, title=f"熱 予測 ({th_lam:.0f} m)")
    s_th_fit.marker = Marker(symbol=None)
    s_th_fit.graphicalProperties.line.solidFill = BLUE
    s_th_fit.graphicalProperties.line.width = 18000
    s_mev_fit = Series(Reference(ws, min_col=8, min_row=28, max_row=fit_last), x_fit, title=f"MeV 予測 ({mev_lam:.0f} m)")
    s_mev_fit.marker = Marker(symbol=None)
    s_mev_fit.graphicalProperties.line.solidFill = RED
    s_mev_fit.graphicalProperties.line.width = 18000

    chart.series.append(s_th)
    chart.series.append(s_mev)
    chart.series.append(s_th_fit)
    chart.series.append(s_mev_fit)
    ws.add_chart(chart, "B42")

    ws["B61"] = "標高差の仮定：KEK = 0 m、筑波山 = 877 m、白根山 = 2000 m（昨年の平均自由行程計算に合わせる）"
    ws["B61"].font = Font(size=10, color=GRAY)

    for col, w in zip("ABCDEFGHI", [3, 22, 16, 22, 18, 16, 22, 22, 4]):
        ws.column_dimensions[col].width = w
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[11].height = 28
    ws.row_dimensions[18].height = 28
    ws.sheet_view.showGridLines = False

    out = "/Users/yuto/KEK_summer/02_解析/宇宙線中性子_平均自由行程.xlsx"
    wb.save(out)
    print(f"thermal λ from last-year def: {th_lam:.0f} m  Λ {th_Lam:.1f}")
    print(f"MeV λ same def: {mev_lam:.0f} m  Λ {mev_Lam:.1f}")
    return out


if __name__ == "__main__":
    print("saved", build())
