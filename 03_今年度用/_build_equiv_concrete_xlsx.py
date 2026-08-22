#!/usr/bin/env python3
"""図11「全地点・等価コンクリート」の解析 Excel を生成する。

黄色=入力、水色=Excel数式（ASCIIのみ）。
理論: A = A0*EXP(-x/lambda_cm)、x・lambda_cm とも等価コンクリート [cm]（横軸と同じ）。
"""

from __future__ import annotations

import math
from pathlib import Path

import equiv_shielding as esh

from openpyxl import Workbook, load_workbook
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
DATA = ROOT / "測定_20260818"
OUT = DATA / "等価コンクリート_減衰解析.xlsx"
FIG = DATA / "figures" / "11_全地点_等価コンクリート.png"

GRAY, RED, ORANGE = "666666", "D62728", "E67E22"
PALE, WHITE, HEADER = "F5F5F5", "FFFFFF", "1F4E79"
YELLOW, GREEN, CYAN = "FFF2CC", "E8F5E9", "DDEBF7"

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

A0 = 1.0
LAMBDA_CM = esh.LAMBDA_CONCRETE_CM  # ≈39.2 cm（詳細計算。旧77 cmは誤り）
RHO_C = esh.RHO_CONCRETE
RHO_LOAM = esh.RHO_LOAM
RHO_JOSO = esh.RHO_JOSO
LOAM_MAX = esh.LOAM_MAX_CM
LAMBDA_AIR = 1475.0
SOIL_PROFILE = esh.DEFAULT_PROFILE


def _x_from_layers(tc: float, ts: float) -> float:
    """コンクリート + 土から質量厚さ X [g/cm²]（プロファイル準拠）。"""
    return tc * RHO_C + esh.soil_mass_thickness_gcm2(ts, profile=SOIL_PROFILE)


# (地点, t_c, t_soil, CPS, 備考) — X は層厚から算出（Book5のKEKB 117.25は使わない）
_SITES_IN = [
    ("地上", 0.0, 0.0, 0.52514597, "基準（屋外）"),
    ("PF", 105.0, 0.0, 0.25108616, ""),
    ("linac", 150.0, 0.0, 0.06478913, ""),
    ("BT", 60.0, 220.0, 0.11475671, "土はロームのみ（220 cm < 3.5 m）"),
    ("KEKB", 80.0, 670.0, 0.0399324, "ローム3.5 m + 常総2.0 m + 下総1.2 m（Book5の117.25は桁誤り）"),
]
SITES = []
for _label, _tc, _ts, _cps, _note in _SITES_IN:
    _x = _x_from_layers(_tc, _ts)
    SITES.append((_label, _tc, _ts, _x, _cps, _note or f"X={_x:.2f}"))


def style_header(ws, row, cols, start=1, fill=HEADER):
    for c in range(start, start + cols):
        cell = ws.cell(row, c)
        cell.font = Font(bold=True, color=WHITE, size=11)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[row].height = 28


def box(ws, r, c, val=None, fmt=None, fill=None, bold=False, align=center):
    cell = ws.cell(r, c, val)
    cell.border = border
    cell.alignment = align
    cell.font = Font(size=11, bold=bold)
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    return cell


def setup_axis(axis, title, major, minor=None, nfmt="General", amin=None, amax=None):
    axis.title = title
    axis.delete = False
    axis.majorTickMark = "out"
    axis.minorTickMark = "out" if minor else "none"
    axis.tickLblPos = "nextTo"
    axis.majorUnit = major
    if minor:
        axis.minorUnit = minor
    axis.numFmt = nfmt
    if amin is not None:
        axis.scaling.min = amin
    if amax is not None:
        axis.scaling.max = amax
    axis.majorGridlines = ChartLines(
        spPr=GraphicalProperties(ln=LineProperties(solidFill="D9D9D9", w=9525))
    )
    axis.spPr = GraphicalProperties(ln=LineProperties(solidFill="000000", w=15875))


def theory(x_cm: float) -> float:
    """A = A0 * exp(-x/λ)、x は等価コンクリート [cm]。"""
    return A0 * math.exp(-x_cm / LAMBDA_CM)


def _fix_xlsx_relationships(path: Path) -> None:
    """openpyxl が書く Target=\"/xl/...\" 絶対パスを相対パスに直す（Excelが開けない対策）。"""
    import io
    import re
    import zipfile

    buf = io.BytesIO(path.read_bytes())
    out = io.BytesIO()
    with zipfile.ZipFile(buf, "r") as zin, zipfile.ZipFile(out, "w", zipfile.ZIP_DEFLATED) as zout:
        for info in zin.infolist():
            data = zin.read(info.filename)
            if info.filename.endswith(".rels"):
                text = data.decode("utf-8")

                def repl(m: re.Match[str]) -> str:
                    target = m.group(1)
                    if not target.startswith("/xl/"):
                        return m.group(0)
                    # workbook.xml.rels → xl/foo
                    if info.filename == "xl/_rels/workbook.xml.rels":
                        rel = target[len("/xl/") :]  # worksheets/sheet1.xml
                        return f'Target="{rel}"'
                    # worksheets/_rels → ../drawings/...
                    if "/worksheets/_rels/" in info.filename:
                        rel = "../" + target[len("/xl/") :]
                        return f'Target="{rel}"'
                    # drawings/_rels → ../charts/...
                    if "/drawings/_rels/" in info.filename:
                        rel = "../" + target[len("/xl/") :]
                        return f'Target="{rel}"'
                    return f'Target="{target.lstrip("/")}"'

                text2 = re.sub(r'Target="([^"]+)"', repl, text)
                data = text2.encode("utf-8")
            zout.writestr(info, data)
    path.write_bytes(out.getvalue())


def build() -> Path:
    wb = Workbook()

    # ---- 1. 概要（最初に作る。シート順を崩さない）----
    wso = wb.active
    wso.title = "概要"
    wso["A1"] = "図11 等価コンクリート減衰 — 解析ブック"
    wso["A1"].font = Font(size=18, bold=True, color=HEADER)
    wso.merge_cells("A1:F1")
    wso["A2"] = "黄色=入力 / 水色=数式。グラフは「減衰曲線」シート上部。"
    wso["A2"].font = Font(size=11, color=GRAY)

    wso["A4"] = "主結果（測定点シートへの参照）"
    wso["A4"].font = Font(size=13, bold=True)
    for c, h in enumerate(["地点", "X", "t_eq[cm]", "相対", "理論", "残差"], 1):
        wso.cell(5, c, h)
    style_header(wso, 5, 6)

    # 測定点は後で作るので、参照行番号を先に決める
    first = 5
    for i, site in enumerate(SITES):
        src = first + i
        dest = 6 + i
        fill = GREEN if site[0] == "地上" else None
        box(wso, dest, 1, f"=測定点!A{src}", fill=fill, align=left)
        box(wso, dest, 2, f"=測定点!D{src}", "0.00", fill=fill)
        box(wso, dest, 3, f"=測定点!G{src}", "0.0", fill=fill)
        box(wso, dest, 4, f"=測定点!F{src}", "0.000", fill=fill)
        box(wso, dest, 5, f"=測定点!H{src}", "0.000", fill=fill)
        box(wso, dest, 6, f"=測定点!I{src}", "0.000", fill=fill)

    wso["A12"] = "開けない原因だったもの: 説明列に λ を含む不正数式 (=A0*EXP(-x/λ)) が入っていた。"
    wso["A12"].font = Font(size=10, color=ORANGE)
    wso.merge_cells("A12:F12")
    wso["A13"] = f"PNG: {FIG.name}"
    wso["A13"].font = Font(size=10, color=GRAY)

    for col, w in enumerate([10, 10, 12, 10, 10, 10], 1):
        wso.column_dimensions[get_column_letter(col)].width = w
    wso.sheet_view.showGridLines = False

    # ---- 2. パラメータ ----
    wsp = wb.create_sheet("パラメータ")
    wsp["A1"] = "定数（黄色=入力。他シートの数式が参照）"
    wsp["A1"].font = Font(size=16, bold=True, color=HEADER)
    wsp["A2"] = "水色=数式。説明は文字列（先頭を = にしない）。"
    wsp["A2"].font = Font(size=11, color=GRAY)

    for c, h in enumerate(["記号", "値", "単位", "説明"], 1):
        wsp.cell(4, c, h)
    style_header(wsp, 4, 4)

    # B5=A0, B6=lambda_cm, B7=lambda_m, B8=rho_c, ...
    rows_p = [
        (5, "A0", A0, "-", "理論の規格化"),
        (6, "lambda_cm", LAMBDA_CM, "cm", "A=A0*EXP(-x/lambda_cm), x is t_eq [cm]"),
        (7, "lambda_m", "=B6/100", "m", "lambda_cm / 100 (=λ_c from detailed MFP)"),
        (8, "rho_c", RHO_C, "g/cm3", "concrete density"),
        (9, "rho_loam", RHO_LOAM, "g/cm3", "Kanto loam"),
        (10, "rho_joso", RHO_JOSO, "g/cm3", "Joso clay"),
        (11, "loam_max", LOAM_MAX, "cm", "loam / joso boundary"),
        (12, "lambda_air", LAMBDA_AIR, "m", "air (reference)"),
        (13, "CPS0", f"=測定点!E{first}", "1/s", "ground CPS (linked)"),
    ]
    for r, sym, val, unit, desc in rows_p:
        box(wsp, r, 1, sym, fill=PALE, bold=True)
        is_f = isinstance(val, str) and val.startswith("=")
        cell = box(wsp, r, 2, val, fill=CYAN if is_f else YELLOW)
        if r == 13:
            cell.number_format = "0.00000000"
        elif isinstance(val, float):
            cell.number_format = "0.00" if val < 10 else "0.0"
        elif is_f and r == 7:
            cell.number_format = "0.00"
        box(wsp, r, 3, unit)
        box(wsp, r, 4, desc, align=left)

    wsp["A15"] = "数式の書き方（文字列・コピー用）"
    wsp["A15"].font = Font(size=12, bold=True)
    for i, t in enumerate(
        [
            "相対: =E5/パラメータ!$B$13",
            "t_eq: =D5/パラメータ!$B$8",
            "理論: =パラメータ!$B$5*EXP(-G5/パラメータ!$B$6)   ※ G5=t_eq[cm], B6=lambda_cm",
            "曲線: =パラメータ!$B$5*EXP(-A5/パラメータ!$B$6)   ※ A5=t_eq[cm]",
        ],
        16,
    ):
        wsp.cell(i, 1, t).font = Font(size=11, name="Menlo")
        wsp.merge_cells(start_row=i, start_column=1, end_row=i, end_column=4)

    for col, w in enumerate([12, 14, 10, 48], 1):
        wsp.column_dimensions[get_column_letter(col)].width = w
    wsp.sheet_view.showGridLines = False

    # ---- 3. 測定点 ----
    ws = wb.create_sheet("測定点")
    ws["A1"] = "測定点（黄色=入力、水色=数式）"
    ws["A1"].font = Font(size=16, bold=True, color=HEADER)
    ws["A2"] = "理論セル例: =パラメータ!$B$5*EXP(-G5/パラメータ!$B$6)  （G=t_eq[cm], B6=lambda=λ_c cm）"
    ws["A2"].font = Font(size=10, color=GRAY)

    headers = [
        "地点",
        "t_c [cm]",
        "t_soil [cm]",
        "X [g/cm2]",
        "CPS",
        "相対",
        "t_eq [cm]",
        "理論A",
        "相対-理論",
        "備考",
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(4, c, h)
    style_header(ws, 4, len(headers))

    for i, (label, tc, ts, _x, cps, note) in enumerate(SITES):
        r = first + i
        fill = GREEN if label == "地上" else None
        box(ws, r, 1, label, fill=fill, align=left, bold=(label == "地上"))
        box(ws, r, 2, tc, "0.0", fill=YELLOW)
        box(ws, r, 3, ts, "0.0", fill=YELLOW)
        # X は層厚から数式（ローム≤4 m / 以深は常総）
        fx_x = (
            f"=B{r}*パラメータ!$B$8"
            f"+MIN(C{r},パラメータ!$B$11)*パラメータ!$B$9"
            f"+MAX(0,C{r}-パラメータ!$B$11)*パラメータ!$B$10"
        )
        box(ws, r, 4, fx_x, "0.00", fill=CYAN)
        box(ws, r, 5, cps, "0.00000000", fill=YELLOW)
        box(ws, r, 6, f"=E{r}/パラメータ!$B$13", "0.000000", fill=CYAN)
        box(ws, r, 7, f"=D{r}/パラメータ!$B$8", "0.0", fill=CYAN)
        # 理論: A0*EXP(-t_eq_cm / lambda_cm)
        box(ws, r, 8, f"=パラメータ!$B$5*EXP(-G{r}/パラメータ!$B$6)", "0.000000", fill=CYAN)
        box(ws, r, 9, f"=F{r}-H{r}", "0.000000", fill=CYAN)
        box(ws, r, 10, note, align=left)

    last = first + len(SITES) - 1

    ws[f"A{last + 2}"] = "密度（パラメータ参照）: コンクリ2.3 / ローム1.35 / 常総1.65、ローム上限400 cm"
    ws[f"A{last + 2}"].font = Font(size=11, color=GRAY)
    ws.merge_cells(start_row=last + 2, start_column=1, end_row=last + 2, end_column=6)

    chk = last + 3
    ws[f"A{chk}"] = "手計算チェック"
    ws[f"A{chk}"].font = Font(size=12, bold=True)
    box(ws, chk + 1, 1, "EXP(-150/λ_c)", fill=PALE, bold=True)
    box(ws, chk + 1, 2, "=パラメータ!$B$5*EXP(-150/パラメータ!$B$6)", "0.000000", fill=CYAN)
    box(ws, chk + 1, 3, "linac理論", fill=PALE)
    box(ws, chk + 1, 4, f"=H{first + 2}", "0.000000", fill=CYAN)

    for col, w in enumerate([10, 10, 12, 12, 12, 12, 12, 10, 12, 12, 32], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.sheet_view.showGridLines = False

    # ---- 4. 減衰曲線 + グラフ（日本語・各点解説）----
    wsc = wb.create_sheet("減衰曲線")
    wsc["A1"] = "理論曲線と測定点（日本語グラフ）"
    wsc["A1"].font = Font(size=16, bold=True, color=HEADER)
    wsc["A2"] = "緑列=グラフ用数値 / 水色=数式。各点の解説は右表とグラフ凡例を参照。"
    wsc["A2"].font = Font(size=10, color=GRAY)

    for c, h in enumerate(["t_eq[cm]", "理論(数式)", "理論(値)"], 1):
        wsc.cell(4, c, h)
    style_header(wsc, 4, 3)

    x_max = max(s[3] / RHO_C for s in SITES) * 1.04
    n_c = 50
    for i in range(n_c + 1):
        r = 5 + i
        x_cm = x_max * i / n_c
        box(wsc, r, 1, round(x_cm, 6), "0.0", fill=YELLOW)
        # A = A0*EXP(-t_eq_cm / lambda_cm)
        box(wsc, r, 2, f"=パラメータ!$B$5*EXP(-A{r}/パラメータ!$B$6)", "0.000000", fill=CYAN)
        box(wsc, r, 3, theory(x_cm), "0.000000", fill=GREEN)

    # 測定点 + ラベル列（グラフの点名・解説用）
    explanations = {
        "地上": "基準（屋外）。相対=1に規格化",
        "PF": "コンクリート105cm。理論より高い（屋内γ等）",
        "linac": "コンクリート150cm。理論曲線に近い",
        "BT": "コンクリ60+土220cm。理論より高い",
        "KEKB": "コンクリ80+ローム400+常総270cm。厚いが相対は理論より高い",
    }

    wsc["F4"] = "測定点データ"
    wsc["F4"].font = Font(size=12, bold=True)
    for c, h in enumerate(
        ["地点", "x[cm]", "相対(数式)", "相対(値)", "理論値", "ラベル", "解説"], 6
    ):
        wsc.cell(4, c, h)
    style_header(wsc, 4, 7, start=6)

    cps0 = SITES[0][4]
    for i, site in enumerate(SITES):
        src = first + i
        r = 5 + i
        label, _tc, _ts, x, cps, _note = site
        x_eq = x / RHO_C
        y = cps / cps0
        y_th = theory(x_eq)
        label_txt = f"{label} {y:.3f}"
        box(wsc, r, 6, label, fill=GREEN if label == "地上" else None, align=left)
        box(wsc, r, 7, x_eq, "0.0", fill=GREEN)
        box(wsc, r, 8, f"=測定点!F{src}", "0.0000", fill=CYAN)
        box(wsc, r, 9, y, "0.0000", fill=GREEN)
        box(wsc, r, 10, y_th, "0.0000", fill=GREEN)
        box(wsc, r, 11, label_txt, align=left, fill=PALE)
        box(wsc, r, 12, explanations[label], align=left)

    chart = ScatterChart()
    chart.title = "相対CPSと理論減衰 A=A0·EXP(-x/λ_c)（x [cm]）"
    chart.height = 14
    chart.width = 16
    chart.legend.position = "b"
    chart.style = 10
    setup_axis(
        chart.x_axis,
        "等価コンクリート厚さ [cm]",
        50,
        25,
        "0",
        0,
        math.ceil(x_max / 50) * 50,
    )
    setup_axis(chart.y_axis, "相対 CPS（地上 = 1）", 0.1, 0.05, "0.0", 0, 1.2)

    s_th = Series(
        Reference(wsc, min_col=3, min_row=5, max_row=5 + n_c),
        Reference(wsc, min_col=1, min_row=5, max_row=5 + n_c),
        title="理論 A=A0·EXP(-x/λ_c)",
    )
    s_th.graphicalProperties.line.solidFill = GRAY
    s_th.graphicalProperties.line.width = 25000
    s_th.marker = Marker(symbol=None)
    chart.series.append(s_th)

    # 各地点: 1点ずつのシリーズ（凡例=地点名）
    point_colors = {
        "地上": "222222",
        "PF": "1F77B4",
        "linac": "D62728",
        "BT": "2CA02C",
        "KEKB": "9467BD",
    }
    for i, site in enumerate(SITES):
        r = 5 + i
        label = site[0]
        s = Series(
            Reference(wsc, min_col=9, min_row=r, max_row=r),
            Reference(wsc, min_col=7, min_row=r, max_row=r),
            title=label,
        )
        s.graphicalProperties.line.noFill = True
        color = point_colors.get(label, RED)
        s.marker = Marker(symbol="circle", size=12)
        s.marker.graphicalProperties.solidFill = color
        s.marker.graphicalProperties.line.solidFill = color
        chart.series.append(s)

    # グラフは右側（理論データ列と重ならない）
    wsc.add_chart(chart, "N4")

    # 解説表は理論曲線データ（行5〜5+n_c）の下に置く。途中に挟むと #VALUE! になる
    note_row = 5 + n_c + 2  # = 57
    wsc.cell(note_row, 1, "各点の解説").font = Font(size=13, bold=True, color=HEADER)
    for c, h in enumerate(["地点", "t_eq [cm]", "相対", "理論", "解説"], 1):
        wsc.cell(note_row + 1, c, h)
    style_header(wsc, note_row + 1, 5)
    for i, site in enumerate(SITES):
        r = 5 + i
        dest = note_row + 2 + i
        box(wsc, dest, 1, f"=F{r}", align=left)
        box(wsc, dest, 2, f"=G{r}", "0.0")
        box(wsc, dest, 3, f"=I{r}", "0.000")
        box(wsc, dest, 4, f"=J{r}", "0.000")
        box(wsc, dest, 5, f"=L{r}", align=left)

    wsc.cell(note_row + 8, 1, "読み方").font = Font(size=12, bold=True)
    wsc.cell(
        note_row + 9,
        1,
        "灰色の曲線が理論減衰。色付き点が測定。linacは理論に近く、PF/BT/KEKBは理論より高い（単純遮蔽だけでは説明しきれない成分がある）。",
    ).font = Font(size=10, color=GRAY)
    wsc.merge_cells(start_row=note_row + 9, start_column=1, end_row=note_row + 9, end_column=5)

    for col, w in enumerate([12, 12, 12, 3, 3, 10, 10, 10, 10, 10, 14, 42], 1):
        wsc.column_dimensions[get_column_letter(col)].width = w
    wsc.column_dimensions["N"].width = 14
    wsc.sheet_view.showGridLines = False

    # ---- 5. 解釈 ----
    wsi = wb.create_sheet("解釈")
    wsi["A1"] = "トラブルシュート"
    wsi["A1"].font = Font(size=16, bold=True, color=HEADER)
    for i, t in enumerate(
        [
            "グラフは「減衰曲線」シート。タイトル・軸・凡例は日本語。各地点は色分けされ凡例に名前が出る。",
            "各点の解説はグラフ下の表（地点・相対・理論・解説）を参照。",
            "本物の数式は水色セル（例: 測定点!H5 =パラメータ!$B$5*EXP(-G5/パラメータ!$B$6)）。",
            "PNG図: figures/11_全地点_等価コンクリート.png。理論は A=A0·exp(-x/λ_c)、xは等価コンクリート厚[cm]。",
        ],
        3,
    ):
        wsi.cell(i, 1, "・" + t).font = Font(size=11)
        wsi.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)
    wsi.column_dimensions["A"].width = 100
    wsi.sheet_view.showGridLines = False

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    _fix_xlsx_relationships(OUT)
    return OUT


def regenerate_figure() -> None:
    import importlib.util

    spec = importlib.util.spec_from_file_location("plot_mca", ROOT / "_plot_mca.py")
    mod = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(mod)
    mod.fig_all_sites_equiv_concrete()


if __name__ == "__main__":
    regenerate_figure()
    path = build()
    print(f"figure: {FIG}")
    print(f"saved {path}")

    wb = load_workbook(path)
    print("sheets:", wb.sheetnames)
    bad = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("=") and any(
                    ch in v for ch in "λρσμ"
                ):
                    bad.append((ws.title, cell.coordinate, v))
    print("bad greek formulas:", len(bad))
    wsc = wb["減衰曲線"]
    print("charts on 減衰曲線:", len(wsc._charts))
    print("測定点!H5:", wb["測定点"]["H5"].value)
    print("減衰曲線!B5:", wsc["B5"].value)
    print("減衰曲線!C5 (numeric):", wsc["C5"].value)
