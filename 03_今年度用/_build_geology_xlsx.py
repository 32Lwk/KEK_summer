#!/usr/bin/env python3
"""KEK地質モデルに基づく地下中性子フラックス予測ブックを生成する。

グラフには目盛り（major/minor tick, 目盛り線, ラベル, 単位）を明示的に付ける。
"""

from __future__ import annotations

import math

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.chart import ScatterChart, BarChart, Reference, Series
from openpyxl.chart.marker import Marker
from openpyxl.chart.axis import ChartLines
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.utils import get_column_letter

BLACK, GRAY, BLUE, RED, ORANGE = "000000", "666666", "1F77B4", "D62728", "E67E22"
PALE, WHITE, HEADER = "F5F5F5", "FFFFFF", "1F4E79"

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

# 昨年と同じ定義：熱中性子・管理棟1階→白根山、一様密度、Δh=2000 m
# λ = Δh / ln(φ₂/φ₁) ≈ 1470 m  ↔  Λ ≈ 147 g/cm²
Lam_th = 147.5
Lam = 99.8  # MeV を同じ定義で換算した参考値
phi0_th = 1.83e-3
phi0 = 7.25e-4
RHO_AIR = 1.00e-3
LAMBDA_AIR_M = Lam_th / RHO_AIR / 100.0  # 1470 m


def style_header(ws, row, cols, start=1, fill=HEADER):
    for c in range(start, start + cols):
        cell = ws.cell(row, c)
        cell.font = Font(bold=True, color=WHITE, size=11)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[row].height = 28


def box(ws, r, c, val=None, fmt=None, fill=None, bold=False, align=center):
    cell = ws.cell(r, c, val)
    cell.border = border
    cell.alignment = align
    cell.font = Font(size=11, bold=bold)
    if fmt:
        cell.number_format = fmt
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    return cell


def setup_axis(axis, title, major, minor=None, nfmt="General",
               amin=None, amax=None):
    """軸に目盛り・ラベル・単位を明示的に設定する。"""
    axis.title = title
    axis.delete = False
    axis.majorTickMark = "out"
    axis.minorTickMark = "out" if minor else "none"
    axis.tickLblPos = "nextTo"
    axis.majorUnit = major
    if minor:
        axis.minorUnit = minor
    axis.numFmt = nfmt
    if amin is not None:
        axis.scaling.min = amin
    if amax is not None:
        axis.scaling.max = amax
    axis.majorGridlines = ChartLines(
        spPr=GraphicalProperties(ln=LineProperties(solidFill="D9D9D9", w=9525))
    )
    if minor:
        axis.minorGridlines = ChartLines(
            spPr=GraphicalProperties(ln=LineProperties(solidFill="EFEFEF", w=6350))
        )
    axis.spPr = GraphicalProperties(ln=LineProperties(solidFill=BLACK, w=15875))


PROFILE = [(3.0, 1.35, "関東ローム"), (2.0, 1.65, "常総粘土"), (1e9, 1.85, "下総層群")]


def lam_m(density, lam=Lam_th):
    """媒質中の平均自由行程 [m]。λ = Λ / ρ / 100"""
    return lam / density / 100.0


def X_of_depth(depth):
    x = 0.0
    rem = depth
    for thick, dens, _name in PROFILE:
        take = min(rem, thick)
        if take <= 0:
            break
        x += dens * take * 100
        rem -= take
    return x


def remaining(depth, lam=Lam_th):
    """深さ d までの残存率 = Π exp(−Δd / λ_layer)"""
    if depth <= 0:
        return 1.0
    att = 1.0
    rem = depth
    for thick, dens, _name in PROFILE:
        take = min(rem, thick)
        if take <= 0:
            break
        att *= math.exp(-take / lam_m(dens, lam))
        rem -= take
    return att


def layer_name(depth):
    if depth <= 3:
        return "関東ローム内"
    if depth <= 5:
        return "常総粘土に到達"
    return "下総層群"


def build():
    wb = Workbook()

    # ===== Sheet 1: 概要 =====
    ws = wb.active
    ws.title = "概要"
    ws["A1"] = "KEK地下測定 — 地質に基づく宇宙線中性子予測"
    ws["A1"].font = Font(size=18, bold=True, color=HEADER)
    ws.merge_cells("A1:G1")
    ws["A2"] = "昨年度9班（熱中性子・管理棟1階基準、空気中 λ ≈ 1470 m）× 筑波台地の浅層地質モデル"
    ws["A2"].font = Font(size=11, color=GRAY)
    ws.merge_cells("A2:G2")

    ws["A4"] = "このブックの内容"
    ws["A4"].font = Font(size=13, bold=True)
    contents = [
        ("地質_KEK", "KEK周辺（筑波台地）の表層〜浅部の地質と密度仮定"),
        ("予測", "層構造を考慮した深さ方向のフラックス予測（主結果・グラフ）"),
        ("測定計画", "地下測定で記録すべき項目と推奨深さ"),
        ("文献_根拠", "地質・減衰長の根拠"),
    ]
    for i, (a, b) in enumerate(contents, 5):
        box(ws, i, 1, a, fill=PALE, bold=True)
        box(ws, i, 2, b, align=left)
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)

    ws["A10"] = "一言でいうと"
    ws["A10"].font = Font(size=13, bold=True)
    for i, t in enumerate([
        "KEKは筑波台地の関東ローム＋常総粘土の上にあり、土の密度が比較的低いため、",
        "空気中の平均自由行程は約 1470 m（熱中性子・管理棟1階→白根山）。土の中では約 1.1 m まで短くなり、深さ2–3 mで熱中性子は地上の数〜十%まで減ると予測。",
        "深さ5 m超では大気起源はほぼ消え、放射起源フロアが見え始める可能性。",
    ], 11):
        ws.cell(i, 1, t).font = Font(size=12)
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=7)

    ws["A15"] = "主予測（熱中性子・管理棟1階基準 φ=1.83E-03、λ_air ≈ 1470 m）"
    ws["A15"].font = Font(size=13, bold=True)
    for c, h in enumerate(
        ["深さ (m)", "その層のλ (m)", "残存率", "予測φ (s⁻¹ cm⁻²)", "通過する主な層"], 1
    ):
        ws.cell(16, c, h)
    style_header(ws, 16, 5)
    for i, d in enumerate([1, 2, 3, 5, 10], 17):
        dens = 1.35 if d <= 3 else (1.65 if d <= 5 else 1.85)
        r = remaining(d, Lam_th)
        vals = [d, lam_m(dens, Lam_th), r, phi0_th * r, layer_name(d)]
        for c, v in enumerate(vals, 1):
            cell = box(ws, i, c, v, align=left if c == 5 else center)
            if c == 2:
                cell.number_format = "0.00"
            if c == 3:
                cell.number_format = "0.0%"
            if c == 4:
                cell.number_format = "0.00E+00"
            if d in (2, 3):
                cell.fill = PatternFill("solid", fgColor="FFF2CC")
    ws["A23"] = "黄色行（2–3 m）が今年度の狙いどころ：まだ測れるが、地上との差がはっきり出る。"
    ws["A23"].font = Font(size=11, color=ORANGE, bold=True)
    ws.merge_cells("A23:G23")

    for col, w in enumerate([14, 22, 12, 20, 28, 12, 12], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.sheet_view.showGridLines = False

    # ===== Sheet 2: 地質 =====
    ws2 = wb.create_sheet("地質_KEK")
    ws2["A1"] = "KEK（つくば市大穂）周辺の地質"
    ws2["A1"].font = Font(size=16, bold=True, color=HEADER)
    ws2["A2"] = "位置: 茨城県つくば市大穂（筑波台地）。標高おおむね 20–40 m。"
    ws2["A2"].font = Font(size=11, color=GRAY)

    ws2["A4"] = "浅層の標準層序（上から下へ）"
    ws2["A4"].font = Font(size=13, bold=True)
    for c, h in enumerate(
        ["深さの目安", "地層名", "岩相・特徴", "湿潤密度 ρ (g/cm³)", "本予測での扱い"], 1
    ):
        ws2.cell(5, c, h)
    style_header(ws2, 5, 5)
    geo_rows = [
        ("表層 0–数m", "関東ローム層（新期ローム）",
         "火山灰質粘性土。含水比高く多孔質。透水性は団粒のため比較的高い。",
         "1.2–1.5（代表 1.35）", "最上層として採用"),
        ("その下 0.15–4 m厚", "常総粘土層",
         "筑波台地でローム直下に広く分布する凝灰質粘土。難透水。層厚は場所で大きく変化。",
         "約 1.65", "ローム下の第2層"),
        ("さらに下位", "下総層群（常総層・木下層など）",
         "更新統の砂・泥互層。つくば地下100 m以浅に広く分布。",
         "砂泥 1.7–2.0（代表 1.85）", "5 m以深の本体"),
        ("深部（参考）", "上総層群〜基盤",
         "より古い堆積岩・変成岩（温泉ボーリング等で確認）。浅い地下実験では非対象。",
         "—", "今回の予測範囲外"),
    ]
    for i, row in enumerate(geo_rows, 6):
        for c, v in enumerate(row, 1):
            box(ws2, i, c, v, align=left if c > 1 else center)
        ws2.row_dimensions[i].height = 48

    ws2["A11"] = "本予測で使う簡易柱状モデル（仮定）"
    ws2["A11"].font = Font(size=13, bold=True)
    for c, h in enumerate(
        ["深度区間 (m)", "地層", "密度 ρ", "平均自由行程 λ (m)", "区間の面密度 ΔX (g/cm²)"], 1
    ):
        ws2.cell(12, c, h)
    style_header(ws2, 12, 5, fill="2E7D32")
    model = [
        ("0 – 3", "関東ローム", 1.35, lam_m(1.35), 1.35 * 3 * 100),
        ("3 – 5", "常総粘土", 1.65, lam_m(1.65), 1.65 * 2 * 100),
        ("5 – 10", "下総層群（砂泥）", 1.85, lam_m(1.85), 1.85 * 5 * 100),
        ("10 – 20", "下総層群（砂泥）", 1.85, lam_m(1.85), 1.85 * 10 * 100),
    ]
    for i, (a, b, c, d, e) in enumerate(model, 13):
        box(ws2, i, 1, a)
        box(ws2, i, 2, b, align=left)
        box(ws2, i, 3, c, "0.00")
        box(ws2, i, 4, d, "0.00")
        box(ws2, i, 5, e, "0")
    ws2["A18"] = f"空気中の平均自由行程 λ_air = {LAMBDA_AIR_M:.0f} m。土中は λ = λ_air × (ρ_air / ρ_土)。"
    ws2["A18"].font = Font(size=11, color=GRAY)
    ws2.merge_cells("A18:E18")

    ws2["A20"] = "注意（地質のばらつき）"
    ws2["A20"].font = Font(size=12, bold=True, color=RED)
    for i, t in enumerate([
        "・常総粘土の厚さは筑波台地で 15–400 cm と幅が大きい。実測地点のボーリングがあれば差し替えること。",
        "・ローム厚も場所依存。KEK構内の掘削記録・建築図があれば最優先で使う。",
        "・コンクリート床・建物自体も覆土として加算する（ρ≈2.3 g/cm³ 程度）。",
        "・確認用: 産総研「地質図Navi」https://gbank.gsj.jp/geonavi/ ／ 関東平野の地下地質・地盤DB",
    ], 21):
        ws2.cell(i, 1, t).font = Font(size=10)
        ws2.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)

    for col, w in enumerate([16, 28, 42, 22, 18], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.sheet_view.showGridLines = False

    # ===== Sheet 3: 予測 =====
    ws3 = wb.create_sheet("予測")
    ws3["A1"] = "地下フラックス予測（KEK地質モデル × 平均自由行程）"
    ws3["A1"].font = Font(size=16, bold=True, color=HEADER)
    ws3["A2"] = ("φ(d) ≈ φ₀ · exp(−d/λ)　空気中 λ ≈ 1470 m（熱中性子・管理棟1階→白根山）　"
                 "土中 λ = 1470 × (0.001 / ρ)　φ₀ = 1.83×10⁻³")
    ws3["A2"].font = Font(size=10, color=GRAY)

    ws3["A4"] = "深さプロファイル（宇宙線大気起源成分の粗い予測）"
    ws3["A4"].font = Font(size=13, bold=True)
    ph = ["深さ d (m)", "その層のλ (m)", "熱 残存率", "熱 予測φ",
          "MeV 残存率(参考)", "MeV 予測φ(参考)", "層の位置", "測定の狙い"]
    for c, h in enumerate(ph, 1):
        ws3.cell(5, c, h)
    style_header(ws3, 5, 8)

    depths = [0, 0.5, 1, 1.5, 2, 2.5, 3, 4, 5, 6, 8, 10]
    tips = {
        0: "地上基準（昨年度相当）",
        1: "差が出始める",
        2: "推奨① はっきり減る",
        3: "推奨② ローム底付近",
        5: "大気起源はほぼ消える",
        10: "放射起源フロア候補",
    }
    for i, d in enumerate(depths, 6):
        dens = 1.35 if d <= 3 else (1.65 if d <= 5 else 1.85)
        r_th = remaining(d, Lam_th)
        r_mev = remaining(d, Lam)
        vals = [d, lam_m(dens, Lam_th), r_th, phi0_th * r_th, r_mev, phi0 * r_mev, layer_name(d), tips.get(d, "")]
        for c, v in enumerate(vals, 1):
            cell = box(ws3, i, c, v, align=left if c >= 7 else center)
            if c == 1:
                cell.number_format = "0.0"
            if c == 2:
                cell.number_format = "0.00"
            if c in (3, 5):
                cell.number_format = "0.0%"
            if c in (4, 6):
                cell.number_format = "0.00E+00"
            if d in (2, 3):
                cell.fill = PatternFill("solid", fgColor="FFF2CC")
            elif d == 5:
                cell.fill = PatternFill("solid", fgColor="FCE4D6")
            elif d >= 10:
                cell.fill = PatternFill("solid", fgColor="E8F5E9")

    last = 5 + len(depths)  # 17
    ws3.cell(last + 1, 1,
             "色: 黄=推奨深さ / オレンジ=大気起源ほぼ消失 / 緑=放射起源が支配しうる深さ").font = Font(size=10, color=GRAY)
    ws3.merge_cells(start_row=last + 1, start_column=1, end_row=last + 1, end_column=8)

    # --- Scatter chart: 深さ vs 残存率（目盛りつき）---
    chart = ScatterChart()
    chart.title = "KEK地質モデル：深さ vs 残存率"
    chart.height = 11
    chart.width = 17
    chart.legend.position = "b"
    setup_axis(chart.x_axis, "深さ d (m)", major=1, minor=0.5, nfmt="0", amin=0, amax=10)
    setup_axis(chart.y_axis, "宇宙線成分の残存率", major=0.1, minor=0.05,
               nfmt="0%", amin=0, amax=1)

    xvals = Reference(ws3, min_col=1, min_row=6, max_row=17)
    s1 = Series(Reference(ws3, min_col=3, min_row=6, max_row=17), xvals, title="熱（λ_air≈1470 m）")
    s1.graphicalProperties.line.solidFill = BLUE
    s1.graphicalProperties.line.width = 25000
    s1.marker = Marker(symbol="circle", size=8)
    s1.marker.graphicalProperties.solidFill = BLUE
    s1.marker.graphicalProperties.line.solidFill = BLUE
    s2 = Series(Reference(ws3, min_col=5, min_row=6, max_row=17), xvals, title="MeV（参考 λ_air≈1000 m）")
    s2.graphicalProperties.line.solidFill = RED
    s2.graphicalProperties.line.width = 25000
    s2.marker = Marker(symbol="triangle", size=8)
    s2.marker.graphicalProperties.solidFill = RED
    s2.marker.graphicalProperties.line.solidFill = RED
    chart.series.append(s1)
    chart.series.append(s2)
    ws3.add_chart(chart, "A20")

    # --- Bar chart: 主要深さの残存率（目盛りつき）---
    for c, h in enumerate(["深さ (m)", "熱残存率 (%)"], 10):
        cell = ws3.cell(5, c, h)
        cell.font = Font(bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=HEADER)
        cell.border = border
        cell.alignment = center
    for i, d in enumerate([0, 1, 2, 3, 5, 10], 6):
        r = 100.0 * remaining(d, Lam_th)
        box(ws3, i, 10, d, "0")
        box(ws3, i, 11, r, "0.0")

    bar = BarChart()
    bar.type = "col"
    bar.title = "主な深さでの熱中性子残存率"
    bar.height = 11
    bar.width = 13
    bar.legend = None
    bar.add_data(Reference(ws3, min_col=11, min_row=5, max_row=11), titles_from_data=True)
    bar.set_categories(Reference(ws3, min_col=10, min_row=6, max_row=11))
    setup_axis(bar.y_axis, "残存率 (%)", major=10, minor=5, nfmt="0", amin=0, amax=100)
    bar.x_axis.title = "深さ (m)"
    bar.x_axis.delete = False
    bar.x_axis.majorTickMark = "out"
    bar.x_axis.tickLblPos = "nextTo"
    bar.gapWidth = 60
    ws3.add_chart(bar, "J20")

    for col, w in enumerate([12, 16, 14, 14, 16, 16, 16, 24, 3, 12, 14], 1):
        ws3.column_dimensions[get_column_letter(col)].width = w
    ws3.sheet_view.showGridLines = False

    # ===== Sheet 4: 測定計画 =====
    ws4 = wb.create_sheet("測定計画")
    ws4["A1"] = "地下測定の計画（KEK向け）"
    ws4["A1"].font = Font(size=16, bold=True, color=HEADER)
    ws4["A3"] = "推奨測定セット"
    ws4["A3"].font = Font(size=13, bold=True)
    for c, h in enumerate(
        ["優先", "地点イメージ", "深さ/覆土", "期待される熱中性子", "ねらい", "測定時間の目安"], 1
    ):
        ws4.cell(4, c, h)
    style_header(ws4, 4, 6)
    plan = [
        ("必須", "屋外または管理棟1階相当の地上", "0 m（基準）", "〜1.83E-03（昨年度1階）", "基準の再現・較正", "昨年度並み"),
        ("必須", "浅い地下／ピット／B1", "約 2–3 m 相当", "地上の約 6–16%", "λ の検証（一番効く）", "地上の 5–20 倍"),
        ("推奨", "やや深い地下", "約 5 m 相当", "地上の <1%", "大気起源の消失確認", "地上の 50–100 倍"),
        ("発展", "十分深い点", "≳10 m 相当", "ほぼフロア", "放射起源の有無", "可能な限り長く"),
    ]
    for i, row in enumerate(plan, 5):
        for c, v in enumerate(row, 1):
            cell = box(ws4, i, c, v, align=left if c > 1 else center)
            if i == 6:
                cell.fill = PatternFill("solid", fgColor="FFF2CC")
        ws4.row_dimensions[i].height = 36

    ws4["A10"] = "記録必須項目（地質を活かすため）"
    ws4["A10"].font = Font(size=13, bold=True)
    for i, t in enumerate([
        "標高・深さ（GLからの深さ）",
        "覆土／床／天井の材質と厚さ（コンクリート厚を必ず）",
        "可能なら構内ボーリング柱状図との対応",
        "気圧・気温（大気側の補正）",
        "熱中性子と MeV を同一条件でペア測定",
        "屋内か屋外か、周囲の壁・水の有無（熱に効く）",
    ], 11):
        ws4.cell(i, 1, f"□ {t}").font = Font(size=11)
        ws4.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)

    ws4["A18"] = "成功の判定"
    ws4["A18"].font = Font(size=13, bold=True)
    ws4["A19"] = "2–3 m 点の熱中性子が、予測（残存率おおよそ 6–16%）のオーダーに入れば、管理棟1階基準の λ≈1470 m とKEK地質モデルが整合。"
    ws4["A20"] = "MeV が熱中性子より速く減れば、昨年どおり MeV の λ が短い（約 1000 m）ことと整合する。"
    ws4["A19"].font = Font(size=11)
    ws4["A20"].font = Font(size=11)
    ws4.merge_cells("A19:F19")
    ws4.merge_cells("A20:F20")

    for col, w in enumerate([10, 28, 16, 22, 22, 18], 1):
        ws4.column_dimensions[get_column_letter(col)].width = w
    ws4.sheet_view.showGridLines = False

    # ===== Sheet 5: 文献 =====
    ws5 = wb.create_sheet("文献_根拠")
    ws5["A1"] = "根拠・参照"
    ws5["A1"].font = Font(size=16, bold=True, color=HEADER)
    ws5["A3"] = "項目"
    ws5["B3"] = "内容"
    style_header(ws5, 3, 2)
    refs = [
        ("平均自由行程", "熱中性子・管理棟1階→白根山: Δh=2000 m、Λ≈147 g/cm²、空気中 λ≈1470 m（一様密度 ρ_air=1.00×10⁻³ g/cm³）。昨年と同じ定義。"),
        ("地形", "KEKは筑波台地（標高おおむね20–40 m）上"),
        ("層序", "表層=関東ローム → 常総粘土 → 下総層群（宇野沢ほか1988、坂田ほか 2018/2024 など）"),
        ("常総粘土", "筑波台地でローム直下に広く分布。層厚15–400 cm。湿潤密度〜1.65 g/cm³"),
        ("関東ローム密度", "湿潤密度 1.2–1.5 g/cm³（代表1.35を使用）"),
        ("地下中性子", "浅い地下で大気起源が消え、深いと放射起源・ミュオン二次が支配（神岡・低レベル測定の文献）"),
        ("地図", "産総研 地質図Navi https://gbank.gsj.jp/geonavi/"),
        ("限界", "本予測は宇宙線大気起源の指数減衰のみ。放射起源フロアやミュオン二次は別途見積が必要。"),
    ]
    for i, (a, b) in enumerate(refs, 4):
        box(ws5, i, 1, a, fill=PALE, bold=True, align=left)
        box(ws5, i, 2, b, align=left)
        ws5.row_dimensions[i].height = 32
    ws5.column_dimensions["A"].width = 16
    ws5.column_dimensions["B"].width = 90
    ws5.sheet_view.showGridLines = False

    out = "/Users/yuto/KEK_summer/03_今年度用/KEK地下測定_地質予測.xlsx"
    wb.save(out)
    return out


if __name__ == "__main__":
    print("saved", build())
