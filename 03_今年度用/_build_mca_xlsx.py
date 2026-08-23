#!/usr/bin/env python3
"""今年度 MCA（Amptek MCA8000D）スペクトルを CSV と Excel に変換する。

raw/*.mca と USB 上の未登録 .mca を自動検出する。
解析窓は 2 系統:
  1) peak_roi … シリアル別共通窓（1715: 314–366, 2162: 350–408）+ 両側帯 NET
  2) wall_191_764 … ³He(n,p)³H 壁効果帯 191–764 keV（ピーク=764 keV 線形校正）+ 右側帯 NET
ファイル内 <<ROI>> は使わない。
"""

from __future__ import annotations

import argparse
import csv
import shutil
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import Reference, ScatterChart, Series
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.marker import Marker
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from mca_common import (
    analyze_roi,
    analyze_wall_window,
    discover_raw_mca,
    discover_usb_mca,
    infer_serial,
    make_id,
    make_label,
    parse_mca,
)

ROOT = Path(__file__).resolve().parent
OUT = ROOT / "測定_20260818"
RAW = OUT / "raw"
TABLES = OUT / "tables"

BLACK, BLUE, RED, GREEN = "000000", "1F77B4", "D62728", "2CA02C"
PURPLE, BROWN, PINK = "9467BD", "8C564B", "E377C2"
LINE_COLORS = [BLUE, RED, GREEN, PURPLE, BROWN, PINK]
PALE, WHITE, HEADER, YELLOW = "F5F5F5", "FFFFFF", "1F4E79", "FFF2CC"

thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

KNOWN_META = {
    "D1_20260818_1552_管理棟2階.mca": {
        "屋内_屋外": "屋内",
        "標高_m": 30,
        "メモ": "短時間。検出器 D1（SN 1715）。ch0はADCオーバーフロー。ROIは共通窓 314–366。",
    },
    "D1_20260818_1730_linac.mca": {
        "屋内_屋外": "屋内",
        "標高_m": 30,
        "メモ": "linac。検出器 D1（SN 1715）。ch0はADCオーバーフロー。ROIは共通窓。",
    },
    "D1_20260819_0832_管理棟2階.mca": {
        "屋内_屋外": "屋内",
        "標高_m": 30,
        "メモ": "終夜。検出器 D1（SN 1715）。ROIは共通窓。",
    },
    "D1_20260819_1344_管理棟1階.mca": {
        "屋内_屋外": "屋内",
        "標高_m": 30,
        "メモ": "検出器 D1（SN 1715）。ROIは共通窓。",
    },
    "d1_20260819_1520_管理棟2階.mca": {
        "屋内_屋外": "屋内",
        "標高_m": 30,
        "メモ": "管理棟2階・約21 h。検出器 d1（SN 2162）。ポリエチレンの緩衝材なし。旧名の放射線棟2階は誤記。CPS差は d 検出器の応答。ch0なし。ROIは共通窓。",
    },
    "D1_20260819_1530_地上.mca": {
        "屋内_屋外": "屋外",
        "標高_m": 30,
        "メモ": "地上。検出器 D1（SN 1715）。ROIは共通窓。",
    },
    "D1_20260819_1854_放射線棟BT.mca": {
        "屋内_屋外": "屋内",
        "標高_m": 30,
        "メモ": "放射線棟 BT。検出器 D1（SN 1715）。ROIは共通窓。",
    },
    "d2_20260819_1859_放射線棟BT.mca": {
        "屋内_屋外": "屋内",
        "標高_m": 30,
        "メモ": "放射線棟 BT。検出器 d2（SN 2162）。CPS差は d 検出器の応答。ROIは共通窓。",
    },
}


def infer_meta(filename: str) -> dict:
    if filename in KNOWN_META:
        return dict(KNOWN_META[filename])
    outdoor = ("地上" in filename) or ("屋外" in filename)
    return {
        "屋内_屋外": "屋外" if outdoor else "屋内",
        "標高_m": 30,
        "メモ": "raw/ から自動検出。ROIはシリアル別共通窓。",
    }


def unique_id(stem: str, used: set[str]) -> str:
    base = make_id(stem)
    if base not in used:
        used.add(base)
        return base
    i = 2
    while f"{base}_{i}" in used:
        i += 1
    uid = f"{base}_{i}"
    used.add(uid)
    return uid


def summarize(run: dict, meta: dict) -> dict:
    counts = meta["counts"]
    live = float(meta["LIVE_TIME"])
    start = meta["START_TIME"]
    date = "2026-08-18"
    try:
        st = datetime.strptime(start, "%m/%d/%Y %H:%M:%S")
        en = st + timedelta(seconds=float(meta["REAL_TIME"]))
        date = st.strftime("%Y-%m-%d") if st.date() == en.date() else f"{st.strftime('%Y-%m-%d')}〜{en.strftime('%m-%d')}"
    except ValueError:
        pass
    real = float(meta["REAL_TIME"])
    total = sum(counts)
    roi = analyze_roi(counts, serial=infer_serial(run["filename"], meta.get("serial", "")))
    wall = analyze_wall_window(counts, serial=infer_serial(run["filename"], meta.get("serial", "")))
    roi_lo, roi_hi, roi_peak = roi.roi_lo, roi.roi_hi, roi.roi_peak
    roi_sum, roi_bg, roi_net_n, roi_net_err = roi.gross, roi.bg, roi.net, roi.err
    ch0 = counts[0]
    ch1_20 = sum(counts[1:21])
    ch21_149 = sum(counts[21:150])
    ch451 = sum(counts[451:])
    excl0 = total - ch0
    dead = 1.0 - live / real if real else 0.0
    peak_ch = max(range(1, min(150, len(counts))), key=lambda i: counts[i])
    return {
        **run,
        "日付": date,
        "開始時刻": meta["START_TIME"],
        "live_s": live,
        "real_s": real,
        "dead_pct": 100.0 * dead,
        "n_channels": meta["n_channels"],
        "データ数": total,
        "測定時間_s": live,
        "計数率_s-1": total / live,
        "計数率_ch0除く_s-1": excl0 / live,
        "ch0": ch0,
        "ch0_cps": ch0 / live,
        "ch1_20": ch1_20,
        "ch1_20_cps": ch1_20 / live,
        "ch21_149": ch21_149,
        "ch21_149_cps": ch21_149 / live,
        "roi_lo": roi_lo,
        "roi_hi": roi_hi,
        "roi_peak": roi_peak,
        "roi_range": f"{roi_lo}–{roi_hi}",
        "roi_auto_lo": roi.roi_auto_lo,
        "roi_auto_hi": roi.roi_auto_hi,
        "roi_auto_range": f"{roi.roi_auto_lo}–{roi.roi_auto_hi}",
        "roi_counts": roi_sum,
        "roi_cps": roi_sum / live,
        "roi_bg_counts": roi_bg,
        "roi_bg_cps": roi_bg / live,
        "roi_net_counts": roi_net_n,
        "roi_net_cps": roi_net_n / live,
        "roi_net_cps_err": roi_net_err / live,
        "roi_net_valid": int(roi.net_valid),
        "roi_warning": roi.warning,
        "bg_mode": roi.bg_mode,
        "sb_lo": f"{roi.sb_lo_lo}–{roi.sb_lo_hi}",
        "sb_hi": f"{roi.sb_hi_lo}–{roi.sb_hi_hi}",
        "sb_lo_lo": roi.sb_lo_lo,
        "sb_lo_hi": roi.sb_lo_hi,
        "sb_hi_lo": roi.sb_hi_lo,
        "sb_hi_hi": roi.sb_hi_hi,
        # --- 壁効果窓 191–764 keV ---
        "wall_lo": wall.roi_lo,
        "wall_hi": wall.roi_hi,
        "wall_peak": wall.roi_peak,
        "wall_e_lo_kev": wall.e_lo_kev,
        "wall_e_hi_kev": wall.e_hi_kev,
        "wall_kev_per_ch": wall.kev_per_ch,
        "wall_counts": wall.gross,
        "wall_cps": wall.gross / live,
        "wall_bg_counts": wall.bg,
        "wall_bg_cps": wall.bg / live,
        "wall_net_counts": wall.net,
        "wall_net_cps": wall.net / live,
        "wall_net_cps_err": wall.err / live,
        "wall_net_valid": int(wall.net_valid),
        "wall_gross_cps": wall.gross / live,
        "wall_gross_cps_err": (wall.gross ** 0.5) / live if wall.gross > 0 else 0.0,
        "wall_warning": wall.warning,
        "wall_bg_mode": wall.bg_mode,
        "wall_sb_lo": f"{wall.sb_lo_lo}–{wall.sb_lo_hi}" if wall.sb_lo_hi else "",
        "wall_sb_hi": f"{wall.sb_hi_lo}–{wall.sb_hi_hi}" if wall.sb_hi_hi else "",
        "wall_sb_lo_lo": wall.sb_lo_lo,
        "wall_sb_lo_hi": wall.sb_lo_hi,
        "wall_sb_hi_lo": wall.sb_hi_lo,
        "wall_sb_hi_hi": wall.sb_hi_hi,
        "ch451_511": ch451,
        "peak_ch": peak_ch,
        "peak_counts": counts[peak_ch],
        "装置": meta["device"],
        "シリアル": infer_serial(run["filename"], meta.get("serial", "")),
        "ゲインGAIA": meta["gaia"],
        "ボード温度_C": meta["board_temp_C"],
        "slow_count": meta["slow_count"],
        "領域": "全スペクトル（未較正）",
        "フラックス_s-1_cm-2": "",
        "気圧_hPa": "",
        "気温_C": "",
        "覆土深さ_m": "",
        "覆土材質": "",
        "密度_g_cm-3_推定": "",
        "counts": counts,
    }


def copy_into_raw(src: Path) -> Path:
    dst = RAW / src.name
    if src.resolve() != dst.resolve():
        shutil.copy2(src, dst)
        print(f"copy    : {src} -> {dst.name}")
    return dst


def ingest(extra: list[Path], use_usb: bool) -> list[Path]:
    RAW.mkdir(parents=True, exist_ok=True)
    TABLES.mkdir(parents=True, exist_ok=True)
    if use_usb:
        for src in discover_usb_mca():
            dst = RAW / src.name
            if not dst.exists() or src.stat().st_mtime > dst.stat().st_mtime + 1:
                copy_into_raw(src)
    for src in extra:
        src = src.expanduser().resolve()
        if not src.is_file():
            raise FileNotFoundError(f"MCA が見つかりません: {src}")
        copy_into_raw(src)
    files = discover_raw_mca(RAW)
    if not files:
        raise FileNotFoundError(
            "raw/ に .mca がありません。ファイルを置くか USB を接続して再実行してください。"
        )
    return files


def load_runs(files: list[Path]) -> list[dict]:
    used: set[str] = set()
    rows = []
    for path in files:
        meta = parse_mca(path)
        stem = path.stem
        run = {
            "id": unique_id(stem, used),
            "場所": make_label(stem),
            "filename": path.name,
            **infer_meta(path.name),
        }
        rows.append(summarize(run, meta))
    rows.sort(key=lambda r: r["開始時刻"])
    return rows


def write_csv(path: Path, fieldnames: list[str], rows: list[dict]) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)


def write_tables(rows: list[dict]) -> None:
    rec_fields = [
        "id",
        "日付",
        "場所",
        "標高_m",
        "覆土深さ_m",
        "覆土材質",
        "密度_g_cm-3_推定",
        "屋内_屋外",
        "領域",
        "データ数",
        "測定時間_s",
        "計数率_s-1",
        "計数率_ch0除く_s-1",
        "フラックス_s-1_cm-2",
        "気圧_hPa",
        "気温_C",
        "開始時刻",
        "live_s",
        "real_s",
        "dead_pct",
        "ch0",
        "ch0_cps",
        "ch1_20",
        "ch1_20_cps",
        "ch21_149",
        "ch21_149_cps",
        "roi_lo",
        "roi_hi",
        "roi_peak",
        "roi_auto_lo",
        "roi_auto_hi",
        "roi_counts",
        "roi_cps",
        "roi_bg_counts",
        "roi_bg_cps",
        "roi_net_counts",
        "roi_net_cps",
        "roi_net_cps_err",
        "roi_net_valid",
        "roi_warning",
        "bg_mode",
        "sb_lo",
        "sb_hi",
        "sb_lo_lo",
        "sb_lo_hi",
        "sb_hi_lo",
        "sb_hi_hi",
        "wall_lo",
        "wall_hi",
        "wall_peak",
        "wall_e_lo_kev",
        "wall_e_hi_kev",
        "wall_net_cps",
        "wall_net_cps_err",
        "wall_net_valid",
        "wall_warning",
        "wall_bg_mode",
        "peak_ch",
        "装置",
        "シリアル",
        "ゲインGAIA",
        "ボード温度_C",
        "filename",
        "メモ",
    ]
    write_csv(TABLES / "測定記録.csv", rec_fields, rows)

    net_fields = [
        "id",
        "場所",
        "シリアル",
        "live_s",
        "roi_lo",
        "roi_hi",
        "roi_peak",
        "roi_auto_lo",
        "roi_auto_hi",
        "roi_counts",
        "roi_cps",
        "roi_bg_counts",
        "roi_bg_cps",
        "roi_net_counts",
        "roi_net_cps",
        "roi_net_cps_err",
        "roi_net_valid",
        "roi_warning",
        "bg_mode",
        "sb_lo",
        "sb_hi",
    ]
    write_csv(TABLES / "ROI_NET_CPS.csv", net_fields, rows)

    wall_fields = [
        "id",
        "場所",
        "シリアル",
        "live_s",
        "wall_lo",
        "wall_hi",
        "wall_peak",
        "wall_e_lo_kev",
        "wall_e_hi_kev",
        "wall_kev_per_ch",
        "wall_counts",
        "wall_cps",
        "wall_bg_counts",
        "wall_bg_cps",
        "wall_net_counts",
        "wall_net_cps",
        "wall_net_cps_err",
        "wall_net_valid",
        "wall_warning",
        "wall_bg_mode",
        "wall_sb_hi",
    ]
    write_csv(TABLES / "WALL_NET_CPS_191_764keV.csv", wall_fields, rows)

    n = rows[0]["n_channels"]
    spec_fields = (
        ["channel"]
        + [f"counts_{r['id']}" for r in rows]
        + [f"cps_{r['id']}" for r in rows]
    )
    spec_rows = []
    for ch in range(n):
        rec = {"channel": ch}
        for r in rows:
            rec[f"counts_{r['id']}"] = r["counts"][ch]
            rec[f"cps_{r['id']}"] = r["counts"][ch] / r["live_s"]
        spec_rows.append(rec)
    write_csv(TABLES / "スペクトル.csv", spec_fields, spec_rows)

    band_rows = []
    bands = [
        ("ch0（オーバーフロー）", lambda r: r["counts"][0:1]),
        ("ch1–20", lambda r: r["counts"][1:21]),
        ("ch21–149", lambda r: r["counts"][21:150]),
        ("共通ROI", lambda r: r["counts"][r["roi_lo"] : r["roi_hi"] + 1]),
        ("ch451–511", lambda r: r["counts"][451:]),
        ("全チャンネル", lambda r: r["counts"]),
        ("ch0除く", lambda r: r["counts"][1:]),
    ]
    for label, fn in bands:
        rec = {"帯域": label}
        for r in rows:
            s = sum(fn(r))
            rec[f"counts_{r['id']}"] = s
            rec[f"cps_{r['id']}"] = s / r["live_s"]
            rec[f"pct_{r['id']}"] = 100.0 * s / r["データ数"]
        band_rows.append(rec)
    band_fields = ["帯域"]
    for r in rows:
        band_fields += [f"counts_{r['id']}", f"cps_{r['id']}", f"pct_{r['id']}"]
    write_csv(TABLES / "チャンネル帯.csv", band_fields, band_rows)


def style_header(ws, row: int, n: int, start: int = 1) -> None:
    for c in range(start, start + n):
        cell = ws.cell(row, c)
        cell.font = Font(bold=True, color=WHITE, size=11)
        cell.fill = PatternFill("solid", fgColor=HEADER)
        cell.alignment = center
        cell.border = border
    ws.row_dimensions[row].height = 28


def apply_border(ws, r, c) -> None:
    cell = ws.cell(r, c)
    cell.border = border
    cell.alignment = left if c == 1 else center


def setup_axis(axis, title, major=None, nfmt="General", amin=None, amax=None, log=False):
    axis.title = title
    axis.delete = False
    axis.majorTickMark = "out"
    axis.tickLblPos = "nextTo"
    axis.numFmt = nfmt
    if major is not None:
        axis.majorUnit = major
    if amin is not None:
        axis.scaling.min = amin
    if amax is not None:
        axis.scaling.max = amax
    if log:
        axis.scaling.logBase = 10
    axis.majorGridlines = ChartLines(
        spPr=GraphicalProperties(ln=LineProperties(solidFill="D9D9D9", w=9525))
    )
    axis.spPr = GraphicalProperties(ln=LineProperties(solidFill=BLACK, w=15875))


def build_xlsx(rows: list[dict]) -> Path:
    wb = Workbook()

    ws = wb.active
    ws.title = "測定概要"
    ws["A1"] = "今年度 MCA 測定"
    ws["A1"].font = Font(bold=True, size=16, color=HEADER)
    ws.merge_cells("A1:F1")
    ws["A2"] = (
        "Amptek MCA8000D。raw/*.mca と USB の未登録ファイルを自動検出。"
        "エネルギー較正なし。ROI は SN 別共通窓 + ピーク外側帯 NET（peak±16ch 外の側帯15ch）。"
    )
    ws["A2"].alignment = left
    ws.merge_cells("A2:F2")
    ws.row_dimensions[2].height = 36

    labels = [
        ("日付", "日付"),
        ("場所", "場所"),
        ("開始時刻", "開始時刻"),
        ("Live time [s]", "live_s"),
        ("Real time [s]", "real_s"),
        ("Dead time [%]", "dead_pct"),
        ("総カウント", "データ数"),
        ("全体 計数率 [s⁻¹]", "計数率_s-1"),
        ("ch0除く 計数率 [s⁻¹]", "計数率_ch0除く_s-1"),
        ("共通ROI", "roi_range"),
        ("ROI ピーク ch", "roi_peak"),
        ("参考 自動ROI", "roi_auto_range"),
        ("NET 有効", "roi_net_valid"),
        ("ROI 警告", "roi_warning"),
        ("背景モード", "bg_mode"),
        ("左側帯", "sb_lo"),
        ("右側帯", "sb_hi"),
        ("ROI gross [s⁻¹]", "roi_cps"),
        ("ROI NET [s⁻¹]", "roi_net_cps"),
        ("ROI NET 誤差 [s⁻¹]", "roi_net_cps_err"),
        ("ch1–20 計数率 [s⁻¹]", "ch1_20_cps"),
        ("低chピーク", "peak_ch"),
        ("装置", "装置"),
        ("ゲイン GAIA", "ゲインGAIA"),
        ("ボード温度 [°C]", "ボード温度_C"),
        ("元ファイル", "filename"),
        ("メモ", "メモ"),
    ]
    ws["A4"] = "項目"
    for j, r in enumerate(rows):
        ws.cell(4, 2 + j, r["場所"])
    ncols = 1 + len(rows)
    style_header(ws, 4, ncols)

    for i, (label, key) in enumerate(labels, start=5):
        ws.cell(i, 1, label)
        vals = [r[key] for r in rows]
        for j, v in enumerate(vals):
            ws.cell(i, 2 + j, v)
        for c in range(1, ncols + 1):
            apply_border(ws, i, c)
        if i % 2 == 0:
            for c in range(1, ncols + 1):
                ws.cell(i, c).fill = PatternFill("solid", fgColor=PALE)

    num_rows = {
        "live_s": "0.0",
        "real_s": "0.0",
        "dead_pct": "0.00",
        "計数率_s-1": "0.00",
        "計数率_ch0除く_s-1": "0.00",
        "roi_cps": "0.00000",
        "roi_net_cps": "0.00000",
        "roi_net_cps_err": "0.00000",
        "ch1_20_cps": "0.00",
    }
    for i, (_, key) in enumerate(labels, start=5):
        if key in num_rows:
            for j in range(len(rows)):
                ws.cell(i, 2 + j).number_format = num_rows[key]

    note_row = 6 + len(labels)
    ws.cell(note_row, 1, "注意")
    ws.cell(note_row, 1).font = Font(bold=True, color=HEADER)
    notes = [
        "raw/*.mca を列挙。USB に未登録の .mca があれば raw/ へコピーする。",
        "ch0 が総カウントの約半分。ADC オーバーフローとして解析では除外するのが安全。",
        "ROI は SN 別共通窓 + ピーク外側帯 NET（端点台形ではない）。参考として自動ROIも記録。",
        "フラックス（φ）はエネルギー窓と効率 ε, S が未確定のため空欄。",
        "生データ: 03_今年度用/測定_20260818/raw/　表: tables/*.csv",
    ]
    for i, t in enumerate(notes, start=note_row + 1):
        ws.cell(i, 1, f"• {t}")
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=max(6, ncols))
        ws.cell(i, 1).alignment = left
        ws.row_dimensions[i].height = 20

    ws.column_dimensions["A"].width = 32
    for j in range(len(rows)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 28
    ws.freeze_panes = "A5"

    ws2 = wb.create_sheet("スペクトル")
    n = rows[0]["n_channels"]
    headers = ["channel"]
    for r in rows:
        headers += [f"counts_{r['id']}", f"cps_{r['id']}"]
    for c, h in enumerate(headers, 1):
        ws2.cell(1, c, h)
    style_header(ws2, 1, len(headers))
    for ch in range(n):
        ws2.cell(ch + 2, 1, ch)
        col = 2
        for r in rows:
            cts = r["counts"][ch]
            cps = cts / r["live_s"]
            ws2.cell(ch + 2, col, cts)
            ws2.cell(ch + 2, col + 1, cps)
            ws2.cell(ch + 2, col + 1).number_format = "0.0000"
            col += 2
        for c in range(1, len(headers) + 1):
            ws2.cell(ch + 2, c).border = border
    for c in range(1, len(headers) + 1):
        ws2.column_dimensions[get_column_letter(c)].width = 16
    ws2.column_dimensions["A"].width = 12
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{n + 1}"

    chart1 = ScatterChart()
    chart1.title = "低ch（ch 1–80）"
    chart1.style = 10
    chart1.y_axis.crosses = "min"
    chart1.legend.position = "r"
    chart1.height = 10
    chart1.width = 18
    setup_axis(chart1.x_axis, "チャンネル", major=10, amin=1, amax=80)
    setup_axis(chart1.y_axis, "計数率 [s⁻¹ / ch]", nfmt="0.0")
    x_zoom = Reference(ws2, min_col=1, min_row=3, max_row=82)
    for i, r in enumerate(rows):
        ycol = 3 + 2 * i
        y = Reference(ws2, min_col=ycol, min_row=3, max_row=82)
        s = Series(y, xvalues=x_zoom, title=r["場所"])
        s.graphicalProperties.line.solidFill = LINE_COLORS[i % len(LINE_COLORS)]
        s.graphicalProperties.line.width = 15000
        s.marker = Marker(symbol=None)
        chart1.series.append(s)
    ws.add_chart(chart1, f"A{note_row + len(notes) + 3}")

    chart2 = ScatterChart()
    chart2.title = "全ch（対数）"
    chart2.style = 10
    chart2.legend.position = "r"
    chart2.height = 10
    chart2.width = 18
    setup_axis(chart2.x_axis, "チャンネル", major=50, amin=0, amax=511)
    setup_axis(chart2.y_axis, "計数率 [s⁻¹ / ch]", nfmt="0.000", amin=1e-4, amax=50, log=True)
    x_all = Reference(ws2, min_col=1, min_row=2, max_row=n + 1)
    for i, r in enumerate(rows):
        ycol = 3 + 2 * i
        y = Reference(ws2, min_col=ycol, min_row=2, max_row=n + 1)
        s = Series(y, xvalues=x_all, title=r["場所"])
        s.graphicalProperties.line.solidFill = LINE_COLORS[i % len(LINE_COLORS)]
        s.graphicalProperties.line.width = 12000
        s.marker = Marker(symbol=None)
        chart2.series.append(s)
    ws.add_chart(chart2, f"A{note_row + len(notes) + 22}")

    ws3 = wb.create_sheet("チャンネル帯")
    ws3["A1"] = "チャンネル帯ごとのカウントと計数率"
    ws3["A1"].font = Font(bold=True, size=14, color=HEADER)
    ws3.merge_cells("A1:J1")
    band_headers = ["帯域"]
    for r in rows:
        band_headers += [f"{r['場所']} カウント", f"{r['場所']} [s⁻¹]", f"{r['場所']} [%]"]
    for c, h in enumerate(band_headers, 1):
        ws3.cell(3, c, h)
    style_header(ws3, 3, len(band_headers))

    bands = [
        ("ch0（オーバーフロー）", lambda r: r["counts"][0:1]),
        ("ch1–20", lambda r: r["counts"][1:21]),
        ("ch21–149", lambda r: r["counts"][21:150]),
        ("共通ROI", lambda r: r["counts"][r["roi_lo"] : r["roi_hi"] + 1]),
        ("ch451–511", lambda r: r["counts"][451:]),
        ("全チャンネル", lambda r: r["counts"]),
        ("ch0除く", lambda r: r["counts"][1:]),
    ]
    for i, (label, fn) in enumerate(bands, start=4):
        ws3.cell(i, 1, label)
        col = 2
        for r in rows:
            s = sum(fn(r))
            ws3.cell(i, col, s)
            ws3.cell(i, col + 1, s / r["live_s"])
            ws3.cell(i, col + 2, 100.0 * s / r["データ数"])
            ws3.cell(i, col + 1).number_format = "0.00"
            ws3.cell(i, col + 2).number_format = "0.0"
            col += 3
        for c in range(1, len(band_headers) + 1):
            apply_border(ws3, i, c)
        if label == "ch0除く":
            for c in range(1, len(band_headers) + 1):
                ws3.cell(i, c).fill = PatternFill("solid", fgColor=YELLOW)
    for c in range(1, len(band_headers) + 1):
        ws3.column_dimensions[get_column_letter(c)].width = 18

    xlsx = OUT / "測定_20260818_MCA.xlsx"
    wb.save(xlsx)
    return xlsx


def list_sources(use_usb: bool) -> None:
    print(f"raw     : {RAW}")
    raw = discover_raw_mca(RAW)
    if raw:
        for p in raw:
            print(f"  {p.name}")
    else:
        print("  （なし）")
    print("USB     :")
    if not use_usb:
        print("  （--no-usb）")
        return
    usb = discover_usb_mca()
    if usb:
        for p in usb:
            print(f"  {p}")
    else:
        print("  （.mca なし）")


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="MCA を自動検出して CSV / Excel / 図を更新する",
    )
    p.add_argument(
        "files",
        nargs="*",
        type=Path,
        help="追加する .mca（raw/ へコピーしてから全ファイルを再処理）",
    )
    p.add_argument("--no-usb", action="store_true", help="USB を走査しない")
    p.add_argument("--no-plot", action="store_true", help="図を作らない")
    p.add_argument("--list", action="store_true", help="検出した .mca を表示して終了")
    return p.parse_args(argv)


def main(argv: list[str] | None = None) -> None:
    args = parse_args(argv)
    if args.list:
        list_sources(use_usb=not args.no_usb)
        return
    files = ingest(args.files, use_usb=not args.no_usb)
    rows = load_runs(files)
    write_tables(rows)
    xlsx = build_xlsx(rows)
    rec = TABLES / "測定記録.csv"
    spec = TABLES / "スペクトル.csv"
    print(f"raw     : {RAW}")
    print(f"記録    : {rec}")
    print(f"スペクトル: {spec}")
    print(f"Excel   : {xlsx}")
    if not args.no_plot:
        try:
            from _plot_mca import main as plot_main

            plot_main()
        except Exception as exc:
            print(f"figures skipped: {exc}")
    for r in rows:
        print(
            f"  {r['場所']}: live={r['live_s']:.1f}s  "
            f"ROI={r['roi_range']} (peak {r['roi_peak']})  "
            f"R={r['計数率_s-1']:.2f} cps  "
            f"NET={r['roi_net_cps']:.5f}±{r['roi_net_cps_err']:.5f} /s"
        )


if __name__ == "__main__":
    main()
